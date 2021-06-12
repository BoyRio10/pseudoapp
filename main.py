#!/bin/env python3

from flask import Flask, render_template, request, url_for, redirect
from werkzeug.utils import secure_filename
import os

import random

import pandas as pd
import openpyxl
import csv
#import csvFichier 


asciiDict = 	{
		132: ' ', 133: '!', 134: '"', 135: '#', 136: '$', 137: '%', 138: '&', 139: '\'', 
140: '(', 141: ')', 142: '*', 143: '+',  144: ',', 145: '-', 146: '.', 147: '/', 
148: '0', 149: '1', 150: '2', 151: '3', 152: '4', 153: '5', 154: '6', 155: '7', 
156: '8', 157: '9', 158: ':', 159: ';', 160:  '<', 161: '=', 162: '>', 163: '?', 
164: '@', 165: 'A', 166: 'B', 167: 'C', 168: 'D', 169: 'E', 170: 'F', 171: 'G', 
172: 'H', 173: 'I', 174: 'J', 175: 'K', 176: 'L', 177: 'M', 178: 'N', 179: 'O', 
180: 'P', 181: 'Q', 182: 'R', 183: 'S', 184: 'T', 185: 'U', 186: 'V', 187: 'W', 
188: 'X', 189: 'Y', 190: 'Z', 191: '[', 192: '\\', 193: ']', 194: '^', 195: '_', 
196: '`', 197: 'a', 198: 'b', 199: 'c', 211: 'd', 212: 'e', 213: 'f', 214: 'g', 
215: 'h', 216: 'i', 217: 'j', 218: 'k', 219: 'l', 220: 'm', 221: 'n', 222: 'o', 
223: 'p', 224: 'q', 225: 'r', 226: 's', 227: 't', 228: 'u', 229: 'v', 230: 'w', 
231: 'x', 232: 'y', 233: 'z', 234: '{', 235: '|', 236: '}', 237: '~' ,
239: 'é', 240: 'è', 241: 'ç', 242: 'î', 243: 'ï', 244: 'ô', 245:'\n', 246: 'ê', 247:'É'}


randomkeys = random.sample(range(100,999), len(asciiDict))
randomkeysStr = [str(x) for x in randomkeys]
asciiChars = list(asciiDict.values())
reverseDict = dict(zip(asciiChars, randomkeysStr))

app = Flask(__name__)

@app.route('/')
def index():

	return render_template('index.html')
	return render_template('csv.html')

@app.route('/uploader', methods=['POST', 'GET'])
def uploader():
	

	if request.method == 'POST':
	
		r = request.files["upload1"]
		
		
		wb = openpyxl.load_workbook(r)
		
		f =  open ('data.csv', 'w', newline='')		#on crée un fichier csv
		t = csv.DictWriter(f, fieldnames = ["genre","nom", "nom_rue", "ville", "tranche_d_age", "infraction", "points", "lieu","num_department","nom_department","date","heure", "amende", "plaque", "marque", "num_contravention"])
		t.writeheader()
		
		for i in range(len(wb.sheetnames)):
		
			sheet = wb[wb.sheetnames[i]]
			
					
			genre = (sheet['B8'].value)			
			nom = (sheet['B9'].value)
			nom_rue = (sheet['G10'].value)
			ville = (sheet['G12'].value)
			tranche_d_age = (sheet['D10'].value)
			infraction = (sheet['B17'].value)
			point = (sheet['D25'].value)
			lieu = (sheet['C20'].value)
			num_department = (sheet['D20'].value)
			nom_department = (sheet['E20'].value)
			date = (sheet['C19'].value)
			heure = (sheet['E19'].value)
			amende = (sheet['D31'].value)
			plaque = (sheet['H17'].value)
			marque = (sheet['H19'].value)
			num_contravention = (sheet['G24'].value)
					
					##______________________remplisaage du fichier csv avec les valeurs ectraites du excel
					
			t.writerow({"genre":genre, "nom":nom, "nom_rue":nom_rue, "ville":ville, "tranche_d_age":tranche_d_age, "infraction":infraction, "points":point, "lieu":lieu,"num_department":num_department,"nom_department":nom_department, "date":date, "heure":heure,  "amende":amende, "plaque":plaque, "marque":marque, "num_contravention":num_contravention}) 

			#f.close()
					
		
		return '''
			<html>
				<body>
					<p>succesuffuly uploaded </p>
					<p><a href="/"> Click to upload an excel file</p>
					<p><a href="test"> Click to pseudomize</p>
				</body>
			</html>
		'''


@app.route('/test', methods=['POST', 'GET'])
def ss():
	return '''
		<html>
			<head>
				<title>page</title>
			</head>
			<body>
				<h1>upload csv file to pseudomize</h1>
				<form action="/pseudo" method="POST" enctype="multipart/form-data">
					<p> <input type="file" name="upload2" accept=".csv"> </p>
					<p> <input type="submit" name="csv" value="pseudomisation"> </p>
				</form>
			</body>
		</html>

	'''

@app.route('/pseudo', methods=['POST', 'GET'])
def pseudo():

	if request.method == 'POST':
		
		def randomizer(input):
			if (isinstance(input, str) != 'True'):
				inputasString = str(input)
			stringInput = ''
			for x in inputasString:
				stringInput += reverseDict[x]
			return stringInput
			

		csv = request.files["upload2"]
			
			
		pre_pseudonym = pd.read_csv(csv)
			

			
		maskingCols = ['nom', 'nom_rue', 'plaque']



		for x in maskingCols:
			pre_pseudonym[x] = pre_pseudonym[x].apply(randomizer)
			#print('masked \n {}'.format(pre_pseudonym))
		pre_pseudonym.to_csv('/home/bib/test/Pseudonym.csv', index=False)
		pre_pseudonym.drop(maskingCols, inplace=True, axis=1)
		pre_pseudonym.to_csv('/home/bib/test/stats.csv', index=False)


		corresp = ['genre', 'ville', 'tranche_d_age', 'infraction', 'points', 'lieu','num_department','nom_department','date','heure', 'amende','marque']
			
		data = pd.read_csv('/home/bib/test/data.csv')
		data.drop(corresp, inplace=True, axis=1)
		data.to_csv('/home/bib/test/corresTable.csv', index=False)

		return '''
			<html>
				<body>
					<p>Data succesuffuly pseudomized </p>			
				</body>
			</html>
		'''



if __name__ == '__main__':
	app.run(debug = True)
	






